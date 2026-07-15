"""
Reconcile the reconstructed `visit_id` against the source AppInsights
`session_id` — answers "shouldn't session_id already group views into visits?".

Both columns PARTITION the page views into groups. This script compares the two
partitions on the store parquet and prints a verdict:

  1. COUNTS        distinct session_id vs visit_id vs person_id, views/group.
  2. SESSION SIZE  how many page views a source session actually spans. If (almost)
                   every session has exactly ONE view, the SDK is minting a fresh
                   session_id per navigation — it is NOT grouping anything.
  3. WITHIN A VISIT how many DISTINCT source session_ids appear inside one
                   reconstructed visit. ~= views/visit  => session_id resets every
                   view (useless); ~= 1 => session_id already captures the visit.
  4. SESSION HEALTH does a source session ever straddle >30 min or >1 person?
                   (a working session must be one person, one sitting.)
  5. AGREEMENT     on the sessions that DO span >1 view, do those views land in a
                   single visit_id and a single person_id? (high => the two agree
                   where session_id is usable at all.)

On top of the console verdict it writes an XLSX analysis (see --xlsx) that answers
the business question "can we measure Avg. time per page on session_id?" — time on
page is only measurable where a page view has a NEXT view inside the SAME group, so
the grouping decides the metric. The workbook compares official session_id against
the reconstructed visit_id on measurability, dwell distribution and per-page effect.

Read-only. Works on the corp store and on the demo store.

Usage (from the SiteOwnerDashboard project root):
    python scripts/reconcile_visit_session.py
    python scripts/reconcile_visit_session.py --store output/site_pageviews.parquet
    python scripts/reconcile_visit_session.py --xlsx output/my_analysis.xlsx --top 50
    python scripts/reconcile_visit_session.py --no-xlsx        # console only
"""

from __future__ import annotations

import argparse
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

PROJECT_DIR = Path(__file__).resolve().parents[1]

# Same rules the pipeline applies in process_site_pageviews.derive_person_visit():
# a dwell gap above the cap is a tab left open, not reading time.
TOS_CAP_SEC = 30 * 60
SESSION_GAP_MIN = 30

# --- corporate XLSX styles (brand style guide §7) ---------------------------
BLACK = "FF000000"
TITLE_FONT = Font(name="Arial", bold=True, color=BLACK, size=12)
SUB_FONT = Font(name="Arial", color="FF7A7870", size=9)
HEADER_FONT = Font(name="Arial", bold=True, color=BLACK, size=9)
HEADER_BORDER = Border(top=Side(style="thin", color=BLACK),
                       bottom=Side(style="thin", color=BLACK))
HEADER_ALIGN_L = Alignment(horizontal="left", vertical="top", wrap_text=True)
HEADER_ALIGN_R = Alignment(horizontal="right", vertical="top", wrap_text=True)
CELL_FONT = Font(name="Arial", size=10, color=BLACK)
CELL_FONT_BOLD = Font(name="Arial", size=10, color=BLACK, bold=True)
CELL_ALIGN_L = Alignment(horizontal="left", vertical="top", wrap_text=True)
CELL_ALIGN_R = Alignment(horizontal="right", vertical="top", wrap_text=True)
ALT_ROW_FILL = PatternFill(start_color="FFECEBE4", end_color="FFECEBE4", fill_type="solid")
FOOT_BORDER = Border(bottom=Side(style="thin", color=BLACK))
TOTAL_BORDER = Border(top=Side(style="medium", color=BLACK),
                      bottom=Side(style="medium", color=BLACK))
# RAG is data-driven status here (the verdict per finding), which is its allowed use.
STATUS_FONT = {
    "OK": Font(name="Arial", size=10, color="FF6F7A1A", bold=True),
    "Warning": Font(name="Arial", size=10, color="FFE4A911", bold=True),
    "Blocker": Font(name="Arial", size=10, color="FFBD000C", bold=True),
}
NUM_KINDS = {"int": "#,##0", "sec": "#,##0.0", "pct": "0.0%", "ratio": "#,##0.00"}


def pct(n, d):
    return f"{(100.0 * n / d):.1f}%" if d else "n/a"


def fmt_dur(sec) -> str:
    """Mirror the dashboard's fmtDur() so the workbook shows what users see."""
    if sec is None or (isinstance(sec, float) and np.isnan(sec)):
        return "–"
    sec = int(round(sec))
    m, s = divmod(sec, 60)
    return f"{m}m {s}s" if m > 0 else f"{s}s"


# --- XLSX writer -----------------------------------------------------------

def write_title(ws, title: str, subtitle: str, width_cols: int) -> None:
    ws.cell(row=1, column=1, value=title).font = TITLE_FONT
    ws.cell(row=2, column=1, value=subtitle).font = SUB_FONT
    for r in (1, 2):
        ws.cell(row=r, column=1).alignment = CELL_ALIGN_L


def write_note(ws, row: int, label: str, text: str, span: int) -> int:
    """A wrapped prose block (executive read) above a table."""
    ws.cell(row=row, column=1, value=label).font = CELL_FONT_BOLD
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=span)
    cell = ws.cell(row=row + 1, column=1, value=text)
    cell.font = CELL_FONT
    cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
    ws.row_dimensions[row + 1].height = max(30, 13 * (1 + len(text) // 110))
    return row + 3


def write_table(ws, columns, rows, start_row: int = 1) -> int:
    """columns: list of dicts {label, key, kind}; kind in text|status|int|sec|pct|ratio|auto.
    'auto' takes its kind from the row's '_fmt'. Returns the row after the table."""
    hdr = start_row
    ws.row_dimensions[hdr].height = 20
    for j, col in enumerate(columns, 1):
        cell = ws.cell(row=hdr, column=j, value=col["label"])
        cell.font = HEADER_FONT
        cell.alignment = HEADER_ALIGN_R if col["kind"] in NUM_KINDS else HEADER_ALIGN_L
        cell.border = HEADER_BORDER

    first_key = columns[0]["key"]
    widths = [len(c["label"]) for c in columns]
    n = len(rows)
    for i, row in enumerate(rows):
        ri = hdr + 1 + i
        is_total = str(row.get(first_key, "")).strip() in ("Total", "Subtotal")
        is_alt = (i % 2 == 1) and not is_total
        for j, col in enumerate(columns, 1):
            kind = row.get("_fmt", "text") if col["kind"] == "auto" else col["kind"]
            val = row.get(col["key"])
            cell = ws.cell(row=ri, column=j)
            numeric = kind in NUM_KINDS
            cell.alignment = CELL_ALIGN_R if numeric else CELL_ALIGN_L
            if is_alt:
                cell.fill = ALT_ROW_FILL
            if val is None or (isinstance(val, float) and np.isnan(val)):
                cell.value = "–"                       # nil as en-dash
                cell.font = CELL_FONT
                shown = "–"
            elif numeric:
                cell.value = int(val) if kind == "int" else float(val)
                cell.number_format = NUM_KINDS[kind]
                cell.font = CELL_FONT_BOLD if is_total else CELL_FONT
                shown = f"{val:,.1f}" if kind != "int" else f"{int(val):,}"
            else:
                cell.value = val
                cell.font = (STATUS_FONT.get(str(val), CELL_FONT) if kind == "status"
                             else CELL_FONT_BOLD if is_total else CELL_FONT)
                shown = str(val)
            widths[j - 1] = max(widths[j - 1], min(len(shown), 60))
            if is_total:
                cell.border = TOTAL_BORDER
            elif i == n - 1:
                cell.border = FOOT_BORDER

    last = hdr + n
    ws.auto_filter.ref = f"A{hdr}:{get_column_letter(len(columns))}{last}"
    ws.freeze_panes = ws.cell(row=hdr + 1, column=1)
    for j, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(j)].width = min(w + 2, 50)
    return last + 2


# --- analysis ---------------------------------------------------------------

def time_on_page(df: pd.DataFrame, ts: pd.Series, group_col: str) -> pd.Series:
    """Seconds until the next page view of the SAME group — the only way a dwell
    time exists. NaN where the view is the group's last one (no next view)."""
    order = df[[group_col]].assign(_ts=ts).sort_values([group_col, "_ts"], kind="mergesort")
    nxt = order.groupby(group_col, sort=False)["_ts"].shift(-1)
    tos = (nxt - order["_ts"]).dt.total_seconds()
    return tos.where(tos <= TOS_CAP_SEC).reindex(df.index)


def grouping_stats(df: pd.DataFrame, tos: pd.Series, group_col: str, page_col: str) -> dict:
    n = len(df)
    meas = tos.notna()
    n_meas = int(meas.sum())
    sizes = df.groupby(group_col).size()
    sub_sec = int((tos < 1).sum())
    clean = tos[tos >= 1]
    pages_total = df[page_col].nunique()
    pages_signal = df.loc[meas, page_col].nunique()
    return {
        "groups": int(sizes.size),
        "views_per_group": n / max(sizes.size, 1),
        "singleton_groups": int((sizes == 1).sum()),
        "singleton_view_share": float(sizes[sizes == 1].sum()) / max(n, 1),
        "measurable": n_meas,
        "measurability": n_meas / max(n, 1),
        "avg_sec": float(tos.mean()) if n_meas else None,
        "median_sec": float(tos.median()) if n_meas else None,
        "sub_second": sub_sec,
        "sub_second_share": (sub_sec / n_meas) if n_meas else None,
        "avg_sec_clean": float(clean.mean()) if len(clean) else None,
        "pages_total": pages_total,
        "pages_signal": pages_signal,
        "pages_signal_share": pages_signal / max(pages_total, 1),
    }


DWELL_BUCKETS = [0, 1, 5, 15, 30, 60, 120, 300, 600, TOS_CAP_SEC + 1]
DWELL_LABELS = ["< 1s (double-fire artifact)", "1–5s", "5–15s", "15–30s", "30–60s",
                "1–2 min", "2–5 min", "5–10 min", "10–30 min"]
SIZE_BUCKETS = [(1, 1, "1 view (no dwell time measurable)"), (2, 2, "2 views"),
                (3, 3, "3 views"), (4, 5, "4–5 views"), (6, 10, "6–10 views"),
                (11, 20, "11–20 views"), (21, 10**9, "21+ views")]


def build_xlsx(df: pd.DataFrame, path: Path, store: Path, top: int) -> dict:
    ts = pd.to_datetime(df["timestamp"], errors="coerce")
    page_col = "page_key" if "page_key" in df.columns else "page_id"
    n = len(df)

    tos_s = time_on_page(df, ts, "session_id")
    tos_v = time_on_page(df, ts, "visit_id")
    s = grouping_stats(df, tos_s, "session_id", page_col)
    v = grouping_stats(df, tos_v, "visit_id", page_col)

    # session health / agreement on the sessions that can be judged at all
    sess_sizes = df.groupby("session_id").size()
    multi_ids = sess_sizes[sess_sizes > 1].index
    ms = df[df["session_id"].isin(multi_ids)]
    if len(ms):
        g = ms.assign(_ts=ts.reindex(ms.index)).groupby("session_id")
        span = (g["_ts"].max() - g["_ts"].min()).dt.total_seconds() / 60
        over_30_share = float((span > 30).mean())
        multi_person_share = float((g["person_id"].nunique() > 1).mean())
        same_visit_share = float((g["visit_id"].nunique() == 1).mean())
    else:
        over_30_share = multi_person_share = same_visit_share = None

    # --- thresholds -> data-driven verdict ---------------------------------
    usable = s["measurability"] >= 0.60 and (s["sub_second_share"] or 0) <= 0.30
    artifact_dominated = (s["sub_second_share"] or 0) > 0.50 and s["measurability"] < 0.20
    gain = (v["measurability"] - s["measurability"])

    if usable:
        headline = (
            f"YES — session_id carries a time signal: {s['measurability']:.1%} of page views have a "
            f"following view in the same session, so Avg. time per page rests on {s['measurable']:,} "
            f"measured dwell times (avg {fmt_dur(s['avg_sec'])}, median {fmt_dur(s['median_sec'])}). "
            f"The reconstructed visit_id measures {v['measurability']:.1%} of views "
            f"(avg {fmt_dur(v['avg_sec'])}) — the two agree; keep Avg. time per page on session_id."
        )
    elif artifact_dominated:
        headline = (
            f"NO — session_id cannot measure Avg. time per page. Dwell time only exists where a view "
            f"has a NEXT view in the SAME group, and only {s['measurability']:.1%} of views do "
            f"({s['measurable']:,} of {n:,}); {s['sub_second_share']:.1%} of those are sub-second "
            f"double-fires (client-side language redirect), not reading time. The average is therefore "
            f"taken over a handful of artifacts and renders as \"0s\", while pages without a single "
            f"pair render as \"–\". The reconstructed visit_id measures {v['measurability']:.1%} of views "
            f"({v['measurable']:,}, avg {fmt_dur(v['avg_sec'])}) — a {gain:.1%}-point gain — because it "
            f"groups a person's views over a 30-min gap instead of trusting the session cookie."
        )
    else:
        headline = (
            f"PARTLY — session_id measures {s['measurability']:.1%} of page views "
            f"({s['measurable']:,} of {n:,}, avg {fmt_dur(s['avg_sec'])}, "
            f"{(s['sub_second_share'] or 0):.1%} of them sub-second artifacts). visit_id measures "
            f"{v['measurability']:.1%} (avg {fmt_dur(v['avg_sec'])}), a {gain:.1%}-point gain. "
            f"Avg. time per page on session_id is directionally readable but under-covered — report it "
            f"with its coverage, or move the metric to visit_id."
        )

    def status(ok: bool, warn: bool = False) -> str:
        return "OK" if ok else ("Warning" if warn else "Blocker")

    findings = [
        {"no": 1, "finding": "Page views per source session",
         "evidence": f"{s['groups']:,} session_id for {n:,} views ({s['views_per_group']:.2f} views/session)",
         "meaning": "Time on page = timestamp of the NEXT view in the same group minus this one. "
                    "At ~1.0 views/session there is no next view, so no dwell time can exist.",
         "status": status(s["views_per_group"] >= 1.5, s["views_per_group"] >= 1.2)},
        {"no": 2, "finding": "Views stuck in single-view sessions",
         "evidence": f"{s['singleton_view_share']:.1%} of views ({s['singleton_groups']:,} sessions of size 1)",
         "meaning": "Every one of these views is permanently excluded from Avg. time per page — "
                    "it is the last view of its session by construction.",
         "status": status(s["singleton_view_share"] <= 0.4, s["singleton_view_share"] <= 0.7)},
        {"no": 3, "finding": "Measurability of Avg. time per page on session_id",
         "evidence": f"{s['measurability']:.1%} of views ({s['measurable']:,} of {n:,})",
         "meaning": "The share of page views the metric is actually computed on. Below ~60% the average "
                    "describes a biased minority (only people who navigated on within one cookie).",
         "status": status(s["measurability"] >= 0.60, s["measurability"] >= 0.30)},
        {"no": 4, "finding": "Sub-second pairs among the measured dwell times (session_id)",
         "evidence": (f"{s['sub_second_share']:.1%} of measured pairs ({s['sub_second']:,})"
                      if s["measurable"] else "no measured pairs at all"),
         "meaning": "Pairs <1s apart are same-instant double-fires (e.g. client-side language redirect), "
                    "not reading time. Where they dominate, AVG(time_on_page_sec) rounds to \"0s\" — "
                    "that is the metric having no signal, not a formatting bug.",
         "status": status((s["sub_second_share"] or 0) <= 0.10, (s["sub_second_share"] or 0) <= 0.30)},
        {"no": 5, "finding": "Avg. time per page as computed on session_id",
         "evidence": (f"{fmt_dur(s['avg_sec'])} avg / {fmt_dur(s['median_sec'])} median"
                      f" · {fmt_dur(s['avg_sec_clean'])} excl. sub-second pairs"),
         "meaning": "What the dashboard shows today. Read it together with finding 3 and 4: the number is "
                    "only as good as the share of views and the share of real pairs behind it.",
         "status": status(usable, not artifact_dominated)},
        {"no": 6, "finding": "Measurability on the reconstructed visit_id (30-min gap per person)",
         "evidence": f"{v['measurability']:.1%} of views ({v['measurable']:,}) · avg {fmt_dur(v['avg_sec'])}",
         "meaning": "The alternative grouping. It ignores the session cookie and joins a person's views "
                    "until they pause for 30 min, so consecutive views stay in one group and their gap "
                    "becomes a real dwell time.",
         "status": status(v["measurability"] >= 0.60, v["measurability"] >= 0.30)},
        {"no": 7, "finding": "Pages with any time signal at all",
         "evidence": f"session_id: {s['pages_signal']:,}/{s['pages_total']:,} ({s['pages_signal_share']:.1%}) · "
                     f"visit_id: {v['pages_signal']:,}/{v['pages_total']:,} ({v['pages_signal_share']:.1%})",
         "meaning": "Pages without a single measured pair render \"–\" in the Avg. time column. This is the "
                    "share of the page inventory the metric can speak about.",
         "status": status(s["pages_signal_share"] >= 0.8, s["pages_signal_share"] >= 0.5)},
        {"no": 8, "finding": "Do multi-view sessions stay within one sitting and one person?",
         "evidence": (f"{len(multi_ids):,} multi-view sessions · {over_30_share:.1%} span >30 min · "
                      f"{multi_person_share:.1%} span >1 person · {same_visit_share:.1%} land in one visit_id"
                      if len(ms) else "none — every session has a single view, nothing to validate"),
         "meaning": "Where session_id groups anything, it should be one person in one sitting. High "
                    "agreement with visit_id means switching the metric to visit_id does not distort the "
                    "sessions that were fine — it only adds the ones session_id lost.",
         "status": status(bool(len(ms)) and (same_visit_share or 0) >= 0.9,
                          bool(len(ms)) and (same_visit_share or 0) >= 0.7)},
    ]

    # --- workbook ----------------------------------------------------------
    generated = pd.Timestamp.now().strftime("%Y-%m-%d %H:%M")
    span_txt = (f"{ts.min():%Y-%m-%d} – {ts.max():%Y-%m-%d}" if ts.notna().any() else "n/a")
    subtitle = (f"Store: {store.name} · {n:,} page views · {span_txt} · "
                f"dwell cap {TOS_CAP_SEC // 60} min · generated {generated}")

    wb = Workbook()

    # Sheet 1 — Verdict
    ws = wb.active
    ws.title = "Verdict"
    write_title(ws, "Is session_id suitable for measuring Avg. time per page?", subtitle, 5)
    r = write_note(ws, 4, "Bottom line", headline, 5)
    write_table(ws, [
        {"label": "#", "key": "no", "kind": "int"},
        {"label": "Finding", "key": "finding", "kind": "text"},
        {"label": "Evidence in this store", "key": "evidence", "kind": "text"},
        {"label": "Why it decides whether Avg. time per page can be measured", "key": "meaning", "kind": "text"},
        {"label": "Verdict", "key": "status", "kind": "status"},
    ], findings, start_row=r)
    ws.column_dimensions["A"].width = 4
    ws.column_dimensions["B"].width = 34
    ws.column_dimensions["C"].width = 42
    ws.column_dimensions["D"].width = 66
    ws.column_dimensions["E"].width = 10

    # Sheet 2 — Time on page, side by side
    ws = wb.create_sheet("Time on Page")
    write_title(ws, "Avg. time per page — official session_id vs reconstructed visit_id", subtitle, 4)
    rows = [
        {"metric": "Page views in store", "s": n, "v": n, "_fmt": "int",
         "note": "Identical universe — only the grouping differs."},
        {"metric": "Distinct groups", "s": s["groups"], "v": v["groups"], "_fmt": "int",
         "note": "More groups = the same journey cut into more pieces."},
        {"metric": "Page views per group", "s": s["views_per_group"], "v": v["views_per_group"], "_fmt": "ratio",
         "note": "A group of 1 view can never yield a dwell time."},
        {"metric": "Groups with exactly 1 view", "s": s["singleton_groups"], "v": v["singleton_groups"], "_fmt": "int",
         "note": "Their views are lost to the metric by construction."},
        {"metric": "Views in 1-view groups", "s": s["singleton_view_share"], "v": v["singleton_view_share"],
         "_fmt": "pct", "note": "Share of traffic that can never be timed."},
        {"metric": "Measurable views (a next view exists)", "s": s["measurable"], "v": v["measurable"],
         "_fmt": "int", "note": "The rows AVG(time_on_page_sec) is actually taken over."},
        {"metric": "Measurability (% of all views)", "s": s["measurability"], "v": v["measurability"],
         "_fmt": "pct", "note": "The coverage of the metric. This is the deciding number."},
        {"metric": "Avg. time per page", "s": s["avg_sec"], "v": v["avg_sec"], "_fmt": "sec",
         "note": "Mean over the measurable views, dwell capped at 30 min."},
        {"metric": "Median time on page", "s": s["median_sec"], "v": v["median_sec"], "_fmt": "sec",
         "note": "Robust counterpart; far below the mean = a few long tails."},
        {"metric": "Sub-second pairs (<1s)", "s": s["sub_second"], "v": v["sub_second"], "_fmt": "int",
         "note": "Double-fires (language redirect), not reading time."},
        {"metric": "Sub-second share of measured pairs", "s": s["sub_second_share"], "v": v["sub_second_share"],
         "_fmt": "pct", "note": "High share => the average collapses to \"0s\"."},
        {"metric": "Avg. time per page excl. sub-second pairs", "s": s["avg_sec_clean"], "v": v["avg_sec_clean"],
         "_fmt": "sec", "note": "The signal left once artifacts are removed."},
        {"metric": "Pages with ≥1 measured view", "s": s["pages_signal"], "v": v["pages_signal"], "_fmt": "int",
         "note": f"Of {s['pages_total']:,} pages in the store."},
        {"metric": "Pages with a time signal (%)", "s": s["pages_signal_share"], "v": v["pages_signal_share"],
         "_fmt": "pct", "note": "The rest render \"–\" in the dashboard."},
    ]
    write_table(ws, [
        {"label": "Metric", "key": "metric", "kind": "text"},
        {"label": "Official session_id\n(dashboard today)", "key": "s", "kind": "auto"},
        {"label": "Reconstructed visit_id\n(30-min gap per person)", "key": "v", "kind": "auto"},
        {"label": "How to read it", "key": "note", "kind": "text"},
    ], rows, start_row=4)
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 22
    ws.column_dimensions["D"].width = 60

    # Sheet 3 — Dwell distribution
    ws = wb.create_sheet("Dwell Distribution")
    write_title(ws, "Where the measured dwell times actually sit", subtitle, 5)
    r = write_note(ws, 4, "Why this matters",
                   "Measurability alone is not enough — the measured pairs must be real reading time. "
                   "Pairs under one second are the same page firing twice (client-side language redirect), "
                   "so a bucket table dominated by \"< 1s\" means Avg. time per page averages artifacts and "
                   "shows \"0s\", no matter how many pairs there are.", 5)
    cut_s = pd.cut(tos_s.dropna(), bins=DWELL_BUCKETS, right=False, labels=DWELL_LABELS)
    cut_v = pd.cut(tos_v.dropna(), bins=DWELL_BUCKETS, right=False, labels=DWELL_LABELS)
    cnt_s, cnt_v = cut_s.value_counts(), cut_v.value_counts()
    tot_s, tot_v = max(int(cnt_s.sum()), 1), max(int(cnt_v.sum()), 1)
    rows = [{"bucket": lab, "s": int(cnt_s.get(lab, 0)), "ss": int(cnt_s.get(lab, 0)) / tot_s,
             "v": int(cnt_v.get(lab, 0)), "vs": int(cnt_v.get(lab, 0)) / tot_v} for lab in DWELL_LABELS]
    rows.append({"bucket": "Total", "s": int(cnt_s.sum()), "ss": 1.0 if int(cnt_s.sum()) else None,
                 "v": int(cnt_v.sum()), "vs": 1.0 if int(cnt_v.sum()) else None})
    write_table(ws, [
        {"label": "Measured dwell time", "key": "bucket", "kind": "text"},
        {"label": "Pairs (session_id)", "key": "s", "kind": "int"},
        {"label": "Share (session_id)", "key": "ss", "kind": "pct"},
        {"label": "Pairs (visit_id)", "key": "v", "kind": "int"},
        {"label": "Share (visit_id)", "key": "vs", "kind": "pct"},
    ], rows, start_row=r)
    ws.column_dimensions["A"].width = 32

    # Sheet 4 — Group size
    ws = wb.create_sheet("Group Size")
    write_title(ws, "How many page views one group holds", subtitle, 5)
    r = write_note(ws, 4, "Why this matters",
                   "The last view of a group has no successor, so a group of n views yields n-1 dwell times. "
                   "The more the grouping fragments, the more views fall out of Avg. time per page — "
                   "1-view groups contribute nothing at all.", 5)
    vis_sizes = df.groupby("visit_id").size()
    rows = []
    for lo, hi, label in SIZE_BUCKETS:
        gs = int(((sess_sizes >= lo) & (sess_sizes <= hi)).sum())
        gv = int(((vis_sizes >= lo) & (vis_sizes <= hi)).sum())
        rows.append({"bucket": label, "s": gs, "sv": int(sess_sizes[(sess_sizes >= lo) & (sess_sizes <= hi)].sum()),
                     "v": gv, "vv": int(vis_sizes[(vis_sizes >= lo) & (vis_sizes <= hi)].sum())})
    rows.append({"bucket": "Total", "s": int(sess_sizes.size), "sv": n, "v": int(vis_sizes.size), "vv": n})
    write_table(ws, [
        {"label": "Group size", "key": "bucket", "kind": "text"},
        {"label": "Sessions (session_id)", "key": "s", "kind": "int"},
        {"label": "Page views in them", "key": "sv", "kind": "int"},
        {"label": "Visits (visit_id)", "key": "v", "kind": "int"},
        {"label": "Page views in them", "key": "vv", "kind": "int"},
    ], rows, start_row=r)
    ws.column_dimensions["A"].width = 32

    # Sheet 5 — Per-page effect
    ws = wb.create_sheet("Pages")
    write_title(ws, f"Top {top} pages — what the dashboard shows vs what is measurable", subtitle, 8)
    r = write_note(ws, 4, "Why this matters",
                   "The per-page view of the same problem: \"Displayed\" reproduces the dashboard's Avg. time "
                   "cell exactly (– when no pair exists, 0s when only sub-second double-fires do). Compare it "
                   "against the coverage columns to see whether a page has a time signal or just a number.", 8)
    d = df.assign(_s=tos_s, _v=tos_v)
    agg = d.groupby(page_col).agg(
        page=("page_name", "first") if "page_name" in df.columns else (page_col, "first"),
        views=(page_col, "size"),
        avg_s=("_s", "mean"), meas_s=("_s", "count"),
        avg_v=("_v", "mean"), meas_v=("_v", "count"),
    ).sort_values("views", ascending=False).head(top)
    rows = []
    for key, rec in agg.iterrows():
        avg_s = None if pd.isna(rec["avg_s"]) else float(rec["avg_s"])
        avg_v = None if pd.isna(rec["avg_v"]) else float(rec["avg_v"])
        rows.append({
            "page": str(rec["page"])[:80], "key": str(key)[:80], "views": int(rec["views"]),
            "disp": fmt_dur(avg_s), "avg_s": avg_s, "cov_s": int(rec["meas_s"]) / max(int(rec["views"]), 1),
            "avg_v": avg_v, "cov_v": int(rec["meas_v"]) / max(int(rec["views"]), 1),
        })
    write_table(ws, [
        {"label": "Page", "key": "page", "kind": "text"},
        {"label": "Page key", "key": "key", "kind": "text"},
        {"label": "Page views", "key": "views", "kind": "int"},
        {"label": "Displayed today (session_id)", "key": "disp", "kind": "text"},
        {"label": "Avg. time, session_id (s)", "key": "avg_s", "kind": "sec"},
        {"label": "Measured views, session_id", "key": "cov_s", "kind": "pct"},
        {"label": "Avg. time, visit_id (s)", "key": "avg_v", "kind": "sec"},
        {"label": "Measured views, visit_id", "key": "cov_v", "kind": "pct"},
    ], rows, start_row=r)
    ws.column_dimensions["A"].width = 46
    ws.column_dimensions["B"].width = 46

    # Sheet 6 — Method
    ws = wb.create_sheet("Method")
    write_title(ws, "How every number in this workbook is computed", subtitle, 3)
    method = [
        {"term": "Time on page", "definition":
            "Seconds from a page view to the person's NEXT page view inside the same group. There is no "
            "browser event for \"left the page\", so this difference is the only available proxy.",
         "rule": "next_view.timestamp − view.timestamp, per group, ordered by timestamp"},
        {"term": "Last view of a group", "definition":
            "Has no successor, therefore no dwell time — it is NULL, not 0, and drops out of the average. "
            "A group of n views yields n−1 dwell times.",
         "rule": "time_on_page_sec IS NULL"},
        {"term": "Dwell cap", "definition":
            "Gaps beyond the cap are a tab left open overnight, not reading time, and would wreck the mean.",
         "rule": f"gap > {TOS_CAP_SEC // 60} min → NULL (same cap as the pipeline)"},
        {"term": "Measurability", "definition":
            "Share of all page views that have a measurable dwell time. The coverage of Avg. time per page — "
            "the deciding number for whether the metric may be published.",
         "rule": "count(time_on_page_sec) / count(*)"},
        {"term": "Official session_id", "definition":
            "The source AppInsights session id and the company standard for counting Visits (set 2026-07-13). "
            "It has a real ~30-min inactivity timeout but is also renewed within a sitting (reloads, new tabs, "
            "blocked cookies), so it over-fragments the journey.",
         "rule": "source column, kept as-is"},
        {"term": "Reconstructed visit_id", "definition":
            "Our own grouping: one person's views form a visit until they pause longer than the gap. Calibrated "
            "against the source's own measured timeout, so it captures the journey the cookie loses.",
         "rule": f"new visit when person changes or gap > {SESSION_GAP_MIN} min"},
        {"term": "person_id", "definition":
            "The real person behind a view: the GPN, else an anonymous device id. Neither session_id nor "
            "user_id identifies a person.",
         "rule": "gpn, else 'anon:' + user_id"},
        {"term": "Sub-second pair", "definition":
            "Two views of the same page under one second apart — a client-side double-fire (e.g. language "
            "redirect). Counted as a measured pair by SQL, but worth no reading time.",
         "rule": "time_on_page_sec < 1"},
        {"term": "Displayed today", "definition":
            "The dashboard's Avg. time cell, reproduced exactly: NULL renders \"–\"; a value under 0.5s rounds "
            "to \"0s\". \"0s\" on a high-traffic page therefore means no time signal, not a formatting bug.",
         "rule": "fmtDur(AVG(time_on_page_sec)) per page"},
        {"term": "Verdict thresholds", "definition":
            "Measurability ≥60% and ≤30% sub-second pairs → session_id carries the metric. Measurability <20% "
            "with >50% sub-second pairs → it cannot. Anything between → publish only with its coverage.",
         "rule": "applied to this store, see the Verdict sheet"},
    ]
    write_table(ws, [
        {"label": "Term", "key": "term", "kind": "text"},
        {"label": "Definition", "key": "definition", "kind": "text"},
        {"label": "Formula / rule", "key": "rule", "kind": "text"},
    ], method, start_row=4)
    ws.column_dimensions["A"].width = 24
    ws.column_dimensions["B"].width = 78
    ws.column_dimensions["C"].width = 46

    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(str(path))
    return {"headline": headline, "session": s, "visit": v}


def main() -> None:
    ap = argparse.ArgumentParser(description="Reconcile visit_id vs source session_id")
    ap.add_argument("--store", default=str(PROJECT_DIR / "output" / "site_pageviews.parquet"))
    ap.add_argument("--examples", type=int, default=5, help="example visits to print")
    ap.add_argument("--xlsx", default=str(PROJECT_DIR / "output" / "session_vs_visit_analysis.xlsx"),
                    help="path of the XLSX analysis workbook")
    ap.add_argument("--no-xlsx", action="store_true", help="console only, skip the workbook")
    ap.add_argument("--top", type=int, default=30, help="pages on the 'Pages' sheet")
    args = ap.parse_args()

    path = Path(args.store)
    if not path.exists():
        raise SystemExit(f"store not found: {path}")
    df = pd.read_parquet(path)
    n = len(df)

    need = ["session_id", "visit_id", "person_id"]
    missing = [c for c in need if c not in df.columns]
    if missing:
        raise SystemExit(f"store is missing {missing} — run process_site_pageviews.py "
                         "(the derived columns are added there). Rebuild with --rebuild.")

    print(f"store: {path}  ({n:,} page views)\n")

    # 1. COUNTS -------------------------------------------------------------
    n_sess = df["session_id"].nunique()
    n_visit = df["visit_id"].nunique()
    n_person = df["person_id"].nunique()
    print("1. COUNTS")
    print(f"   distinct source session_id : {n_sess:,}   ({n/max(n_sess,1):.2f} views/session)")
    print(f"   distinct visit_id (recon.) : {n_visit:,}   ({n/max(n_visit,1):.2f} views/visit)")
    print(f"   distinct person_id         : {n_person:,}   ({n/max(n_person,1):.2f} views/person)")

    # 2. SESSION SIZE -------------------------------------------------------
    sess_sizes = df.groupby("session_id").size()
    one_view_sessions = int((sess_sizes == 1).sum())
    views_in_one = int(sess_sizes[sess_sizes == 1].sum())
    print("\n2. SOURCE SESSION SIZE (page views per source session_id)")
    print(f"   sessions with exactly 1 view : {one_view_sessions:,}/{n_sess:,} "
          f"({pct(one_view_sessions, n_sess)} of sessions)")
    print(f"   page views in 1-view sessions: {views_in_one:,}/{n:,} "
          f"({pct(views_in_one, n)} of views)")
    dist = sess_sizes.value_counts().sort_index()
    head = ", ".join(f"{k}v×{v:,}" for k, v in dist.head(6).items())
    print(f"   size histogram (views×count) : {head}"
          + (" …" if len(dist) > 6 else ""))

    # 3. WITHIN A RECONSTRUCTED VISIT --------------------------------------
    per_visit = df.groupby("visit_id").agg(
        views=("session_id", "size"),
        sess=("session_id", "nunique"),
        persons=("person_id", "nunique"),
    )
    multi = per_visit[per_visit["views"] > 1]
    one_session_share = 0.0
    print("\n3. WITHIN ONE RECONSTRUCTED VISIT")
    print(f"   multi-view visits            : {len(multi):,}/{len(per_visit):,}")
    if len(multi):
        ratio = (multi["sess"] / multi["views"]).mean()
        print(f"   mean distinct session_id / views in those visits: {ratio:.2f}")
        print(f"     ~1.0 => a NEW session_id per page view (session_id useless);")
        print(f"     ~1/views (low) => session_id already captures the visit.")
        one_session = int((multi["sess"] == 1).sum())
        one_session_share = one_session / len(multi)
        print(f"   visits whose views ALL share ONE session_id: {one_session:,}/{len(multi):,} "
              f"({pct(one_session, len(multi))})")

    # 4. SOURCE SESSION HEALTH ---------------------------------------------
    print("\n4. SOURCE SESSION HEALTH (only sessions with >1 view can be judged)")
    multi_sess_ids = sess_sizes[sess_sizes > 1].index
    if len(multi_sess_ids) == 0:
        print("   none — every source session has a single view, so session_id")
        print("   cannot be validated at all. It is effectively per-page-view.")
    else:
        ms = df[df["session_id"].isin(multi_sess_ids)]
        g = ms.groupby("session_id")
        span_min = (g["timestamp"].max() - g["timestamp"].min()).dt.total_seconds() / 60
        over_30 = int((span_min > 30).sum())
        multi_person = int((g["person_id"].nunique() > 1).sum())
        print(f"   multi-view sessions          : {len(multi_sess_ids):,}")
        print(f"   spanning >30 min             : {over_30:,} ({pct(over_30, len(multi_sess_ids))})")
        print(f"   spanning >1 person_id        : {multi_person:,} ({pct(multi_person, len(multi_sess_ids))})")

        # 5. AGREEMENT ------------------------------------------------------
        same_visit = int((g["visit_id"].nunique() == 1).sum())
        same_person = int((g["person_id"].nunique() == 1).sum())
        print("\n5. AGREEMENT (do multi-view sessions land in ONE visit / ONE person?)")
        print(f"   all views in a single visit_id : {same_visit:,}/{len(multi_sess_ids):,} "
              f"({pct(same_visit, len(multi_sess_ids))})")
        print(f"   all views for a single person  : {same_person:,}/{len(multi_sess_ids):,} "
              f"({pct(same_person, len(multi_sess_ids))})")

    # examples --------------------------------------------------------------
    if args.examples and len(multi):
        print(f"\nEXAMPLES ({args.examples} multi-view visits: views | distinct session_id)")
        for vid, r in multi.head(args.examples).iterrows():
            print(f"   {str(vid)[:32]:32s}  views={int(r['views'])}  distinct_session_id={int(r['sess'])}")

    # VERDICT ---------------------------------------------------------------
    per_view_share = views_in_one / max(n, 1)
    print("\nVERDICT")
    if per_view_share > 0.9:
        print("   session_id is minting a fresh id per page view (>90% of views sit in")
        print("   1-view sessions). It does NOT group a visit — the reconstructed")
        print("   visit_id is the correct grouping. Your intuition is right in theory;")
        print("   this AppInsights instance just doesn't persist the session cookie.")
    elif one_session_share > 0.8:
        print("   session_id DOES group views into visits: in "
              f"{pct(int(one_session_share*len(multi)), len(multi))} of multi-view")
        print("   visits every view shares one session_id, and multi-view sessions")
        print("   stay within one person and one sitting. Here session_id and visit_id")
        print("   agree — reconstruction is a safety net, not a correction.")
    else:
        print("   Mixed: session_id groups some visits but resets within others.")
        print("   Reconstruction (visit_id) normalises both cases — keep counting on it.")

    # XLSX ANALYSIS ---------------------------------------------------------
    if not args.no_xlsx:
        out = Path(args.xlsx)
        res = build_xlsx(df, out, path, args.top)
        s, v = res["session"], res["visit"]
        print("\nAVG. TIME PER PAGE — measurability of the two groupings")
        print(f"   session_id : {s['measurability']:>6.1%} of views measurable "
              f"({s['measurable']:,})  avg {fmt_dur(s['avg_sec'])}  "
              f"sub-second pairs {(s['sub_second_share'] or 0):.1%}")
        print(f"   visit_id   : {v['measurability']:>6.1%} of views measurable "
              f"({v['measurable']:,})  avg {fmt_dur(v['avg_sec'])}  "
              f"sub-second pairs {(v['sub_second_share'] or 0):.1%}")
        print(f"\nAnalysis written to {out}")
        print("   sheets: Verdict | Time on Page | Dwell Distribution | Group Size | Pages | Method")


if __name__ == "__main__":
    main()
